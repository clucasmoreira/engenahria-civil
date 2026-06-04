import os
import re
import pandas as pd
from PyPDF2 import PdfReader
from tkinter import Tk, filedialog
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image

# ==== Seleção da pasta ====
root = Tk()
root.withdraw()
pasta = filedialog.askdirectory(title="Selecione a pasta com os PDFs e TXTs")
if not pasta:
    print("❌ Nenhuma pasta selecionada. Encerrando.")
    exit()

# ==== Tamanhos padrão (mm) ====
FORMATOS = {
    "A0": (841, 1189),
    "A1": (594, 841),
    "A0_EXT": (841, 1500),   # A0 Estendido
    "A1_EXT": (594, 1189),   # A1 Estendido
}

TOL = 15  # tolerância em mm

def dentro_tolerancia(valor, alvo):
    """Retorna True se valor estiver dentro da tolerância do alvo"""
    return abs(valor - alvo) <= TOL

def classificar_formato(largura, altura):
    """Classifica o tamanho do PDF como A0, A1, A0 Ext, A1 Ext ou Outro"""
    
    w, h = sorted([largura, altura])  # garante (menor, maior)

    if (dentro_tolerancia(w, FORMATOS["A1"][0]) and
        dentro_tolerancia(h, FORMATOS["A1"][1])):
        return "A1"

    elif (dentro_tolerancia(w, FORMATOS["A0"][0]) and
          dentro_tolerancia(h, FORMATOS["A0"][1])):
        return "A0"

    elif (dentro_tolerancia(w, FORMATOS["A1_EXT"][0]) and
          dentro_tolerancia(h, FORMATOS["A1_EXT"][1])):
        return "A1 Ext"

    elif (dentro_tolerancia(w, FORMATOS["A0_EXT"][0]) and
          dentro_tolerancia(h, FORMATOS["A0_EXT"][1])):
        return "A0 Ext"

    else:
        return f"Outro ({w:.1f}x{h:.1f} mm)"

def ler_txt(caminho_txt):
    """Extrai informações chave,valor do TXT"""
    dados = {}
    encodings = ["utf-8", "latin1", "cp1252"]  # tenta várias codificações
    for enc in encodings:
        try:
            with open(caminho_txt, encoding=enc) as f:
                for linha in f:
                    if "," in linha:
                        chave, valor = linha.strip().split(",", 1)
                        dados[chave.strip()] = valor.strip()
            break  # se deu certo, sai do loop
        except UnicodeDecodeError:
            continue
        except FileNotFoundError:
            break
    return dados

def montar_assunto(dados_txt):

    textos = []

    for campo in [
        "CONTEUDO_LINHA1",
        "CONTEUDO_LINHA2",
        "CONTEUDO_LINHA3"
    ]:

        texto = dados_txt.get(campo, "").strip()

        if texto:
            textos.append(texto)

    if len(textos) == 0:
        return ""

    if len(textos) == 1:
        return textos[0]

    if len(textos) == 2:
        return f"{textos[0]} E {textos[1]}"

    return f"{', '.join(textos[:-1])} E {textos[-1]}"

def extrair_info_pdf(caminho_pdf):
    """Extrai info do PDF e tenta ler TXT correspondente"""
    info = {}

    # Lê TXT associado
    nome_base = os.path.splitext(os.path.basename(caminho_pdf))[0]
    caminho_txt = os.path.join(os.path.dirname(caminho_pdf), nome_base + ".txt")
    dados_txt = ler_txt(caminho_txt)
    info["N°"] = dados_txt.get("NUM_PLANTA", "")

    info["NOME DO ARQUIVO"] = re.sub(
        r"-R\d{2}$",
        "",
        nome_base,
        flags=re.IGNORECASE
    )

    try:
        reader = PdfReader(caminho_pdf)
        page = reader.pages[0]
        box = page.mediabox
        
        # dimensões em pontos (1pt = 25.4 / 72 mm)
        largura_mm = float(box.width) * 25.4 / 72
        altura_mm = float(box.height) * 25.4 / 72

        info["TIPO FOLHA"] = classificar_formato(largura_mm, altura_mm)
    except Exception as e:
        info["Erro_PDF"] = str(e)

    info["ASSUNTO"] = montar_assunto(dados_txt)
    info["R00"] = dados_txt.get("DATA", "")
    info["R01"] = dados_txt.get("DATA_2", "")
    info["R02"] = dados_txt.get("DATA_3", "")
    info["R03"] = dados_txt.get("DATA_4", "")
    info["R04"] = dados_txt.get("DATA_5", "")
    info["R05"] = dados_txt.get("DATA_6", "")
    info["R06"] = dados_txt.get("DATA_7", "")
    info["STATUS"] = ""

    info["NOMEDI"] = dados_txt.get("NOMEDI", "")
    info["TITCLI"] = dados_txt.get("TITCLI", "")
    info["PROJETO"] = dados_txt.get("PROJETO", "")
    info["NOME_PROJETISTA"] = dados_txt.get("NOME_PROJETISTA", "")

    return info

# ==== Loop na pasta ====
dados = []
for arquivo in os.listdir(pasta):
    if arquivo.lower().endswith(".pdf"):
        caminho = os.path.join(pasta, arquivo)
        dados.append(extrair_info_pdf(caminho))

# ==== Exporta para Excel ====
df = pd.DataFrame(dados)

# Informações do cabeçalho
titulo = df["NOMEDI"].iloc[0]
cliente = df["TITCLI"].iloc[0]
projeto = df["PROJETO"].iloc[0]
responsavel = df["NOME_PROJETISTA"].iloc[0]

total_pranchas = int(
    pd.to_numeric(df["N°"], errors="coerce")
    .fillna(0)
    .max()
)

emissao = datetime.now().strftime("%d/%m/%Y")

# Remove colunas usadas apenas para o cabeçalho
df = df.drop(
    columns=[
        "NOMEDI",
        "TITCLI",
        "PROJETO",
        "NOME_PROJETISTA"
    ],
    errors="ignore"
)

# Nome do arquivo
timestamp = datetime.now().strftime("%Y%m%d-%H-%M")

nome_arquivo = f"LISTA_MESTRA_{timestamp}.xlsx"

saida = os.path.join(pasta, nome_arquivo)

# CRIA EXCEL FORMATADO
wb = Workbook()
ws = wb.active

ws.title = "Lista Mestra"
ws.sheet_view.showGridLines = False     # Remove as linhas de grade

# ==== LOGO ====
img_path = os.path.join(os.path.dirname(__file__), "logo.png")

if os.path.exists(img_path):
    img = Image(img_path)

    # opcional: ajustar tamanho
    img.width = 227
    img.height = 91

    ws.add_image(img, "B2")
else:
    print("⚠️ Logo não encontrada:", img_path)

# CABEÇALHO
ws.merge_cells("D2:M2")

ws["D2"] = "LISTAGEM DE PLANTAS"
ws["D2"].font = Font(bold=True)
ws["D2"].alignment = Alignment(horizontal="center")

ws["D3"] = "TÍTULO:"
ws["D3"].font = Font(bold=True)
ws["E3"] = titulo

ws["D4"] = "CLIENTE:"
ws["D4"].font = Font(bold=True)
ws["E4"] = cliente

ws["D5"] = "RESPONSÁVEL:"
ws["D5"].font = Font(bold=True)
ws["E5"] = responsavel

ws["F3"] = "N° PROJETO:"
ws["F3"].font = Font(bold=True)
ws["G3"] = projeto

ws["F4"] = "N° PRANCHAS:"
ws["F4"].font = Font(bold=True)
ws["G4"] = total_pranchas

ws["F5"] = "EMISSÃO:"
ws["F5"].font = Font(bold=True)
ws["G5"] = emissao

# TABELA
linha_inicio = 7

# Cabeçalho da tabela
for col_num, coluna in enumerate(df.columns, start=2):

    celula = ws.cell(
        row=linha_inicio,
        column=col_num
    )

    celula.value = coluna
    celula.font = Font(bold=True)

    celula.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

# Dados
for i, row in enumerate(
    df.values,
    start=linha_inicio + 1
):

    for j, valor in enumerate(
        row,
        start=2
    ):

        ws.cell(
            row=i,
            column=j,
            value=valor
        )

# CENTRALIZA DADOS DA TABELA (exceto ASSUNTO)
for linha in range(linha_inicio + 1, ws.max_row + 1):

    for col in [2, 3, 4, 6, 7, 8, 9, 10, 11, 12, 13]:
        ws.cell(
            linha,
            col
        ).alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

# LARGURAS DAS COLUNAS
larguras = {
    "B": 15,  # N°
    "C": 25,  # NOME DO ARQUIVO
    "D": 15,  # TIPO FOLHA
    "E": 60,  # ASSUNTO
    "F": 15,  # R00
    "G": 15,  # R01
    "H": 15,  # R02
    "I": 15,  # R03
    "J": 15,  # R04
    "K": 15,  # R05
    "L": 15,  # R06
    "M": 15,  # STATUS
}

for coluna, largura in larguras.items():
    ws.column_dimensions[coluna].width = largura

# BORDAS DA TABELA
borda_fina = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for linha in ws.iter_rows(
    min_row=linha_inicio,
    max_row=ws.max_row,
    min_col=2,
    max_col=ws.max_column
):
    for celula in linha:
        celula.border = borda_fina

# SALVA
wb.save(saida)

print()
print("Arquivo criado com sucesso:")
print(saida)